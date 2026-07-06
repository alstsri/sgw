#!/usr/bin/env python3
"""ShopGoodwill menswear hunter — full search library, live API."""
import argparse
import csv
import html
import json
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus

import requests

API_ROOT = "https://buyerapi.shopgoodwill.com/api"
ITEM_ROOT = "https://shopgoodwill.com/item"
USER_AGENT = "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:12.0) Gecko/20100101 Firefox/12.0"

# ---------------------------------------------------------------------------
# BUYER PROFILE — edit this block to configure for a different buyer.
# All sizing filters and scoring logic read from here; search strings in
# SEARCH_GROUPS below should be updated to match when you change sizes.
# ---------------------------------------------------------------------------
BUYER_PROFILE = {
    # --- Shoes ---
    # Primary US sizes accepted. Also accepts UK/EU equivalents within 0.5.
    # For dress shoes (run slim): 7.5 and 8M/8D are both valid.
    "shoe_sizes": ["8", "7.5"],
    # Width codes that are acceptable on a size-8 shoe
    "shoe_widths": ["", "D", "M", "E", "EE"],
    # US sizes to hard-reject (anything at or above this)
    "shoe_reject_above": 8.5,

    # --- Jackets / Tailoring / Outerwear ---
    # US suit/jacket sizes accepted (exact label matches)
    "jacket_sizes_us": ["34S", "34R", "34", "32R", "32"],
    # Italian jacket sizes accepted (IT label on garment)
    "jacket_sizes_it": ["42", "44"],
    # Hard-reject US jacket sizes at or above this number
    "jacket_reject_us_above": 35,
    # Hard-reject Italian jacket sizes at or above this number
    "jacket_reject_it_above": 46,

    # --- Shirts ---
    # Primary neck size (inches). Most shirts must be at or near this.
    "shirt_neck": "15",
    # Neck size accepted ONLY for the exceptional makers listed below.
    "shirt_neck_exception": "15.5",

    # --- Pants ---
    # Waist size in inches. Only pants at this waist are accepted.
    "pants_waist": "28",

    # --- Unstructured / Knitwear / Workwear body size ---
    # Used for pieces without structured sizing (sweaters, chore coats, etc.)
    "body_size": "small",   # "small" | "medium" | "large"
}

# ---------------------------------------------------------------------------
# Patterns derived from BUYER_PROFILE — computed once, used in filters/scoring.
# These update automatically when BUYER_PROFILE changes.
# ---------------------------------------------------------------------------
_shoe_accept_re = re.compile(
    r'(?<![\d.])(?:'
    + '|'.join(
        re.escape(s).replace(r'\.', r'[._\s]?') + r'(?:\s*[dmew]{0,2})?'
        for s in BUYER_PROFILE["shoe_sizes"]
    )
    + r')(?![\d.])',
    re.I,
)
_shoe_reject_above = BUYER_PROFILE["shoe_reject_above"]

_jacket_accept_re = re.compile(
    r'\b(?:'
    + '|'.join(re.escape(s) for s in BUYER_PROFILE["jacket_sizes_us"] + BUYER_PROFILE["jacket_sizes_it"])
    + r'|' + re.escape(BUYER_PROFILE["body_size"])
    + r')\b',
    re.I,
)
_jacket_reject_us_above = BUYER_PROFILE["jacket_reject_us_above"]
_jacket_reject_it_above = BUYER_PROFILE["jacket_reject_it_above"]

_pants_waist = BUYER_PROFILE["pants_waist"]
_shirt_neck = BUYER_PROFILE["shirt_neck"]
_shirt_neck_exception = BUYER_PROFILE["shirt_neck_exception"]
_body_size = BUYER_PROFILE["body_size"]

# ---------------------------------------------------------------------------
# Brand / fabric intelligence
# ---------------------------------------------------------------------------

QUALITY_BRANDS = {
    # shoes — American/Italian
    "alden", "allen edmonds", "red wing", "white's", "wesco", "viberg",
    "carmina", "carlos santos", "meermin", "santoni", "testoni", "bontoni",
    "magnanni", "tod's", "bally", "ferragamo", "stefano bemer", "silvano lattanzi",
    "marsell", "marsèll", "guidi", "officine creative",
    "russell moccasin", "rancourt", "yuketen", "danner",
    # shoes — English
    "crockett", "tricker", "edward green", "john lobb", "gaziano", "church",
    "cheaney", "joseph cheaney", "grenson", "loake", "alfred sargent",
    "george cleverley", "foster & son",
    # shoes — French
    "weston", "paraboot", "berluti", "heschung", "corthay", "aubercy",
    "bowen paris", "septieme largeur", "septième largeur", "jacques demeter",
    # tailoring — Italian
    "loro piana", "brunello", "zegna", "ermenegildo", "canali", "corneliani",
    "brioni", "kiton", "isaia", "belvest", "sartorio", "boglioli", "caruso",
    "ring jacket", "attolini", "cesare attolini", "cantarelli",
    # tailoring — English / Savile Row
    "anderson & sheppard", "anderson sheppard", "henry poole", "huntsman",
    "gieves", "norton & sons", "norton sons", "davies & son", "dege & skinner",
    "richard james", "timothy everest", "ozwald boateng", "chittleborough",
    "margaret howell", "hackett", "cordings", "daks", "aquascutum", "crombie",
    "private white", "belstaff", "barbour", "grenfell", "mackintosh",
    # tailoring — French
    "arnys", "cifonelli", "camps de luca", "smalto", "husbands paris",
    "de bonne facture", "officine generale", "lemaire", "le mont saint michel",
    "hermes", "hermès",
    # tailoring — American
    "oxxford", "hickey freeman", "golden fleece", "j. press", "j press",
    "southwick", "samuelsohn", "jack victor", "paul stuart", "phineas cole",
    "ralph lauren", "purple label", "rrl", "pendleton", "harris tweed",
    "schott", "o'connell",
    # trousers
    "incotex", "zanella", "pt01", "pt torino", "berwich", "rota",
    "marco pescarolo", "bill's khakis", "hertling", "epaulet",
    # shirts
    "turnbull", "asser", "charvet", "eton", "barba", "finamore", "borrelli",
    "fray", "lorenzini", "gitman", "individualized", "mercer", "drake",
    "kamakura", "hilditch", "harvie", "new & lingwood", "emma willis",
    "budd", "figaret", "courtot",
    # loungewear / underwear / base-layer / premium basics
    "hanro", "zimmerli", "derek rose", "calida", "schiesser", "mey",
    "falke", "majestic filatures", "cdlp", "hamilton and hare",
    "luca faloni", "tekla", "dagsmejan", "merz b. schwanen",
    "lady white co", "the white briefs", "lunya", "james perse",
    # knitwear
    "saint james", "armor lux", "sunspel", "drumohr", "inis meain",
    "william lockie", "johnstons of elgin", "johnstons", "n.peal",
    "john smedley", "smedley", "ballantyne", "pringle", "malo",
    "inverallan", "howlin",
    # workwear / heritage outerwear
    "vetra", "le laboureur", "filson", "pointer brand",
    # Japanese
    "kapital", "engineered garments", "needles", "45rpm", "orSlow", "orslow",
    "monitaly", "blue blue japan", "visvim", "comoli", "camoshita",
    "scye", "beams plus", "nonnative", "sage de cret",
    # accessories
    "e. marinella", "marinella", "charvet",
    # blankets
    "faribault", "woolrich", "avoca", "begg & co", "begg",
    # shoes additions
    "vass", "sanders", "quoddy", "esquivel",
}

FABRIC_TERMS = {
    "cashmere", "camel hair", "camelhair", "wool", "harris tweed", "tweed",
    "houndstooth", "hopsack", "flannel", "fresco", "alpaca", "mohair",
    "linen", "silk", "horn button", "shell cordovan", "cordovan", "calfskin",
    "goodyear", "made in england", "made in italy", "made in usa",
    "made in scotland", "made in ireland", "made in japan", "made in france",
    "merino", "moleskin", "moleskin", "sea island", "donegal",
}

MALL_BRANDS = {
    "zara", "h&m", "shein", "forever 21", "puma", "true religion",
    "calvin klein", "cole haan", "banana republic", "express",
}

# Hard-reject brands — never fetched or scored regardless of search string
REJECT_BRANDS = {
    # mid-market / mass shoes
    "clarks", "bass", "g.h. bass", "florsheim", "nunn bush", "rockport",
    "skechers", "steve madden", "aldo", "bostonian", "naturalizer",
    "johnston & murphy", "dexter", "born", "børn",
    # fast fashion / mall apparel
    "stafford", "lands' end", "land's end", "van heusen", "geoffrey beane",
    "arrow", "alfani", "club room", "tasso elba", "dockers", "izod",
    "jos. a. bank", "jos a bank", "express",
    # denim brands that flood size-34 tailoring results
    "wrangler", "brooklyn jeans", "topman",
    # department-store house brands
    "john henry", "croft & barrow", "st. john's bay",
}

# ---------------------------------------------------------------------------
# Search library — complete buyer brief, verbatim
# ---------------------------------------------------------------------------

SEARCH_GROUPS: dict[str, dict[str, Any]] = {
    "shoes": {
        "category_id": 161,
        "category_level": 3,
        "strings": [
            # American / Italian makers
            "Allen Edmonds 8", "Allen Edmonds 8D", "Allen Edmonds 7.5",
            "Alden 8", "Alden 8D", "Alden 7.5",
            "Red Wing 8", "Red Wing 7.5",
            "White's boots 8", "White's boots 7.5",
            "Wesco 8", "Wesco 7.5",
            "Viberg 8", "Viberg 7.5",
            "Carmina 8", "Carmina 7.5",
            "Carlos Santos 8", "Carlos Santos 7.5",
            "Meermin 8", "Meermin 7.5",
            "Santoni 8", "Santoni 7.5",
            "Testoni 8", "Testoni 7.5",
            "Bontoni 8", "Bontoni 7.5",
            "Magnanni 8", "Magnanni 7.5",
            "Tod's 8", "Tod's 7.5",
            "Bally 8", "Bally 7.5",
            "Ferragamo 8", "Ferragamo 7.5",
            "Salvatore Ferragamo 8", "Salvatore Ferragamo 7.5",
            "Stefano Bemer 8", "Stefano Bemer 7.5",
            "Silvano Lattanzi 8", "Silvano Lattanzi 7.5",
            # English makers (US 8 / 7.5 / UK 7 / EU 41)
            "Crockett Jones 8", "Crockett & Jones 8", "Crockett Jones 7.5",
            "Crockett Jones 7", "Crockett & Jones 7",
            "Tricker's 8", "Trickers 8", "Tricker's 7.5",
            "Edward Green 8", "Edward Green 7.5", "Edward Green 7",
            "John Lobb 8", "John Lobb 7.5", "John Lobb 7",
            "Gaziano Girling 8", "Gaziano & Girling 8", "Gaziano Girling 7.5",
            "Church's 8", "Church's 7.5", "Church's 7",
            "Cheaney 8", "Joseph Cheaney 8", "Cheaney 7.5",
            "Grenson 8", "Grenson 7.5",
            "Loake 1880 8", "Loake 1880 7.5",
            "Alfred Sargent 8", "Alfred Sargent 7.5",
            "George Cleverley 8", "George Cleverley 7.5",
            "Foster & Son 8", "Foster & Son 7.5",
            # French makers (US 8 / 7.5 / EU 41 / 40.5)
            "JM Weston 8", "J.M. Weston 8", "Weston Paris 8",
            "JM Weston 7.5", "J.M. Weston 7.5",
            "Paraboot 8", "Paraboot 7.5", "Paraboot 41",
            "Berluti 8", "Berluti 7.5", "Berluti 41",
            "Heschung 8", "Heschung 7.5",
            "Corthay 8", "Corthay 7.5",
            "Aubercy 8", "Aubercy 7.5",
            "Septieme Largeur 8", "Septième Largeur 8",
            # Designer collaborations
            "Ralph Lauren Crockett Jones 8", "Ralph Lauren Edward Green 8",
            # Additional quality makers
            "Vass 8", "Vass 7.5",
            "Sanders 8", "Sanders 7.5",
            "Quoddy 8", "Quoddy 7.5",
            "Visvim 8", "Visvim 7.5",
            "Esquivel 8", "Esquivel 7.5",
            # Italian artisan / American heritage (cross-added from Marissa's makers)
            "Marsell 8", "Marsell 7.5", "Marsèll 8",
            "Guidi 8", "Guidi 7.5",
            "Officine Creative 8", "Officine Creative 7.5",
            "Russell Moccasin 8", "Russell Moccasin 7.5",
            "Rancourt 8", "Rancourt 7.5",
            "Yuketen 8", "Yuketen 7.5",
            # Danner — heritage stitchdown models only (plain "Danner" floods tactical)
            "Danner Mountain Light 8", "Danner Mountain Light 7.5",
            "Danner Bull Run 8", "Danner Bull Run 7.5",
            "Danner Stumptown 8", "Danner Sharptail 8",
        ],
    },
    "tailoring_outerwear": {
        "category_id": 28,
        "category_level": 2,
        "strings": [
            # Italian makers — IT42 (≈ US 32R) primary, IT44 (≈ US 34S) secondary
            "Canali 42", "Canali 44", "Canali 34S",
            "Canali blazer", "Canali cashmere",
            "Corneliani 42", "Corneliani 44",
            "Zegna 42", "Ermenegildo Zegna 42", "Zegna 44",
            "Giorgio Armani 42", "Armani Collezioni 42",
            "Loro Piana 42", "Brunello Cucinelli 42",
            "Brioni 42", "Kiton 42", "Isaia 42",
            "Belvest 42", "Sartorio 42", "Boglioli 42", "Caruso 42",
            "Attolini 42", "Cesare Attolini 42", "Cantarelli 42",
            "Ring Jacket 42",
            # English / London — Savile Row (US 34S / 34R / 32R)
            "Anderson Sheppard 34", "Anderson & Sheppard 34",
            "Henry Poole 34", "Huntsman 34",
            "Gieves Hawkes 34", "Gieves & Hawkes 34",
            "Norton Sons 34", "Norton & Sons 34",
            "Davies & Son 34", "Dege Skinner 34", "Dege & Skinner 34",
            "Richard James 34", "Timothy Everest 34",
            "Ozwald Boateng 34", "Chittleborough Morgan 34",
            # English heritage RTW
            "Margaret Howell men", "Hackett 34", "Hackett London",
            "Cordings", "Cordings Piccadilly",
            "Daks 34", "Aquascutum 34",
            "Burberry vintage 34", "Burberry trench",
            "Mackintosh men", "Grenfell jacket",
            "Crombie overcoat 34", "Crombie coat",
            "Private White VC", "Belstaff 34",
            "Barbour 34", "Drake's London", "Harris Tweed 34",
            # French
            "Arnys Paris", "Arnys",
            "Cifonelli 42", "Cifonelli Paris",
            "Camps de Luca", "Smalto 34", "Francesco Smalto",
            "Husbands Paris", "De Bonne Facture",
            "Officine Generale 34", "Officine Generale men",
            "APC men", "A.P.C. men",
            "Agnes b men", "Agnès b. homme",
            "Lemaire men", "Le Mont Saint Michel",
            "Old England Paris", "Hermes men",
            "Hermes cashmere", "Hermes jacket 34",
            # American
            "Oxxford 34", "Hickey Freeman 34",
            "Brooks Brothers Golden Fleece 34",
            "J Press 34", "J. Press 34",
            "Southwick 34", "Samuelsohn 34", "Jack Victor 34",
            "Paul Stuart 34", "Phineas Cole 34",
            "Ralph Lauren Purple Label 34", "Polo Ralph Lauren 34",
            "RRL 34", "Pendleton 34", "Schott 34", "O'Connell's",
            # Outerwear — fabric & silhouette
            "camel hair 34", "cashmere coat 34", "wool cashmere coat 34",
            "camel hair sport coat 34", "overcoat 34", "topcoat 34",
            "car coat 34", "polo coat 34", "trench coat 34",
            "shearling jacket 34", "sheepskin jacket 34",
            "B-3 jacket 34", "B3 jacket 34", "aviator jacket 34",
            "houndstooth 34", "tweed sport coat 34",
            "flannel suit 34", "fresco suit 34", "hopsack blazer 34",
            "Donegal tweed 34", "Shetland sport coat 34",
            # Japanese soft tailoring
            "Beams Plus 42", "Comoli 42", "Scye 42", "Camoshita 42",
            # Heritage outerwear
            "Gloverall 34",
            # Small / XS catch-all for unstructured pieces
            "Filson wool vest 34", "Filson cruiser 34",
        ],
    },
    "workwear": {
        "category_id": 28,
        "category_level": 2,
        "strings": [
            "French chore jacket", "French workwear",
            "moleskin chore coat", "bleu de travail",
            "French worker jacket", "indigo chore coat", "French work coat",
            "Vetra", "Vetra chore", "Le Laboureur",
            "Filson Mackinaw", "Filson wool vest 34", "Filson cruiser 34",
            "Pointer Brand chore coat", "Lee 91-J",
            "Big Mac chore jacket", "Pendleton wool shirt",
            # Japanese workwear / heritage
            "Kapital small", "Engineered Garments small", "Needles small",
            "45rpm small", "orSlow small", "Monitaly small",
            "Blue Blue Japan small", "Nonnative small",
            # American military heritage
            "CPO jacket small", "N-1 deck jacket small", "M-65 field jacket small",
        ],
    },
    "knitwear": {
        "category_id": 28,
        "category_level": 2,
        "strings": [
            "Saint James wool", "Armor Lux wool", "Sunspel men",
            "Drumohr", "Inis Meain", "William Lockie",
            "Johnstons of Elgin", "N.Peal cashmere",
            "Loro Piana sweater", "Brunello Cucinelli sweater",
            # Additions
            "John Smedley small", "Ballantyne cashmere",
            "Pringle Scotland", "Malo cashmere", "Inverallan small",
        ],
    },
    "accessories": {
        "category_id": 28,
        "category_level": 2,
        "strings": [
            # Pocket squares (ties removed per user request)
            "Drake's pocket square", "Charvet pocket square",
        ],
    },
    "shirts": {
        "category_id": 28,
        "category_level": 2,
        "strings": [
            # Italian makers (neck 15)
            "Brioni shirt 15", "Kiton shirt 15", "Isaia shirt 15",
            "Zegna shirt 15", "Canali shirt 15", "Corneliani shirt 15",
            "Barba Napoli shirt 15", "Finamore shirt 15",
            "Borrelli shirt 15", "Luigi Borrelli shirt 15",
            "Fray shirt 15", "Lorenzini shirt 15", "Truzzi shirt 15",
            # French makers (neck 15)
            "Charvet shirt 15", "Hermes shirt 15",
            "Figaret Paris shirt 15", "Figaret shirt 15", "Courtot shirt 15",
            # English / London (neck 15)
            "Turnbull Asser shirt 15", "Turnbull & Asser shirt 15",
            "Hilditch Key shirt 15", "Hilditch & Key shirt 15",
            "Harvie Hudson shirt 15", "Harvie & Hudson shirt 15",
            "New Lingwood shirt 15", "New & Lingwood shirt 15",
            "Emma Willis shirt 15", "Budd Shirtmakers 15", "Budd shirt 15",
            "Thomas Pink shirt 15", "TM Lewin shirt 15",
            "Jermyn Street shirt 15", "Drake's shirt 15",
            # American & Japanese (neck 15)
            "Eton shirt 15", "Gitman shirt 15", "Gitman Vintage shirt 15",
            "Individualized Shirts 15", "Mercer Sons shirt 15",
            "Mercer & Sons shirt 15", "Kamakura shirt 15",
            # Exceptional-maker 15.5 only
            "Charvet shirt 15.5", "Hermes shirt 15.5",
            "Turnbull Asser shirt 15.5", "Turnbull & Asser shirt 15.5",
            "Brioni shirt 15.5", "Kiton shirt 15.5",
            "Borrelli shirt 15.5", "Finamore shirt 15.5",
            "Barba shirt 15.5", "Fray shirt 15.5",
            # Fabric
            "Sea Island cotton shirt 15", "oxford cloth shirt 15",
            "OCBD shirt 15", "broadcloth shirt 15", "poplin shirt 15",
            "pinpoint shirt 15", "linen shirt 15", "flannel shirt 15",
            "made in England shirt 15", "made in Italy shirt 15",
            "made in USA shirt 15", "made in France shirt 15",
        ],
    },
    "loungewear_basics": {
        # Low-frequency premium underwear / loungewear / base-layer / tee makers.
        # Plain brand searches in Men's Clothing; buyer screens hits manually.
        "category_id": 28,
        "category_level": 2,
        "strings": [
            "Hanro", "Zimmerli", "Zimmerli of Switzerland",
            "Sunspel", "Derek Rose", "Derek Rose London",
            "Calida Switzerland",   # "Calida" alone collides with Spanish "calidad"
            "Schiesser", "Schiesser Revival", "Mey", "Falke",
            "Majestic Filatures", "CDLP", "Hamilton and Hare",
            "Luca Faloni", "Tekla", "Dagsmejan", "Merz b. Schwanen",
            "Lady White Co", "The White Briefs", "Lunya",
            "Turnbull & Asser", "James Perse",
        ],
    },
    "pants": {
        "category_id": 28,
        "category_level": 2,
        "strings": [
            # Italian (28 waist)
            "Zegna 28 trousers", "Canali 28 trousers",
            "Corneliani 28 trousers", "Giorgio Armani 28 trousers",
            "Armani Collezioni 28", "Incotex 28", "Zanella 28",
            "PT01 28", "PT Torino 28", "Berwich 28",
            "Rota 28", "Marco Pescarolo 28",
            # American
            "Ralph Lauren 28 pants", "Ralph Lauren 28 trousers",
            "Polo Ralph Lauren 28", "RRL 28",
            "Brooks Brothers 28 pants", "J Press 28", "J. Press 28",
            "Bill's Khakis 28", "Hertling 28", "Epaulet 28",
            "O'Connell's 28", "Southwick 28",
            # Fabric & origin
            "corduroy 28", "wide wale corduroy 28", "wool trousers 28",
            "flannel trousers 28", "pleated trousers 28",
            "made in Italy trousers 28", "made in USA trousers 28",
            "made in England trousers 28", "made in France trousers 28",
            "made in Japan trousers 28",
        ],
    },
    "blankets": {
        "category_id": 0,
        "category_level": 1,
        "strings": [
            "Pendleton wool blanket", "Faribault wool blanket",
            "Hudson Bay point blanket", "Hudson's Bay point blanket",
            "Woolrich wool blanket", "alpaca blanket",
            "merino wool throw", "cashmere throw", "mohair throw",
            "Scottish wool throw", "Irish wool blanket",
            "English wool blanket", "vintage wool blanket",
            "camp blanket", "made in USA wool blanket",
            "made in Scotland wool throw", "made in Ireland wool blanket",
            "made in England wool blanket", "made in Italy throw",
            "Avoca throw", "Begg & Co", "Johnstons of Elgin throw",
        ],
    },
}

# ---------------------------------------------------------------------------
# Exceptional-maker rule: size 15.5 only accepted for these makers
# ---------------------------------------------------------------------------
EXCEPTIONAL_MAKERS_15_5 = {
    "charvet", "hermes", "hermès", "turnbull", "asser", "brioni", "kiton",
    "borrelli", "finamore", "barba", "fray",
}

# Italian luxury brands where "44" in title means IT44 (≈ US 34–36) — acceptable
ITALIAN_LUXURY = {
    "canali", "corneliani", "zegna", "armani", "brioni", "kiton", "isaia",
    "boglioli", "sartorio", "caruso", "loro piana", "brunello", "cucinelli",
    "belvest", "attolini", "cantarelli",
}


# ---------------------------------------------------------------------------
# Pre-fetch title filter — runs before get_detail(), saves API calls
# ---------------------------------------------------------------------------

def pre_fetch_reject(category: str, title: str) -> tuple[bool, str]:
    """Return (should_skip, reason) based on title alone before fetching detail."""
    t = title.lower()

    # Hard brand reject — applies to all categories
    for brand in REJECT_BRANDS:
        if brand in t:
            return True, f"reject brand: {brand}"

    # ---- SHOES ---------------------------------------------------------------
    if category == "shoes":
        right = bool(_shoe_accept_re.search(t))
        # Build a reject pattern: any size >= shoe_reject_above that isn't an accepted size
        _reject_above = _shoe_reject_above
        wrong = bool(re.search(
            r'\b(?:sz\.?\s*|size\s+)?(?:'
            + '|'.join(
                re.escape(str(n)).replace(r'\.', r'[._\s]?')
                for n in [8.5, 9, 9.5, 10, 10.5, 11, 11.5, 12, 12.5, 13]
                if n >= _reject_above
            )
            + r')\s*[dmew]{0,2}\b', t
        ))
        if wrong and not right:
            return True, f"shoe size outside target ({'/'.join(BUYER_PROFILE['shoe_sizes'])})"

    # ---- SHIRTS & KNITWEAR ---------------------------------------------------
    if category in ("shirts", "knitwear"):
        # Reject XL / XXL / Large / Medium (buyer wears Small / neck 15)
        if re.search(
            r'\b(?:size\s+|sz\.?\s+)?(?:xx?l|2xl|3xl|x-large|xx-large|2x-large|'
            r'2x\b|3x\b|x-large\b)\b', t
        ):
            return True, "size XL/XXL — too large"
        if re.search(r'\b(?:size\s+|sz\.?\s+)(?:large|medium)\b', t):
            return True, "size Large/Medium — too large"
        if re.search(r'\b(?:size\s+|sz\.?\s+)l\b(?!ong|arge|t\.?d)', t):
            return True, "size L — too large"
        if re.search(r'[-–|]\s*(?:size\s+)?(?:large|l)\s*$', t):
            return True, "size L/Large — too large"
        if re.search(r"\bmen'?s\s+(?:size\s+)?(?:l|large|medium|m)\b(?!ong|arge)", t):
            return True, "Men's size L/Large/M/Medium — too large"
        # Standalone "Size M" (not part of "Size M..." like "Size M 15")
        if re.search(r'\b(?:size\s+|sz\.?\s+)m\b(?!\s*\d)', t):
            return True, "size M — too large"

    # ---- WORKWEAR ------------------------------------------------------------
    if category == "workwear":
        if re.search(
            r'\b(?:xx?l|2xl|3xl|x-large|xx-large|2x-large|2x\b|3x\b)\b', t
        ):
            return True, "workwear size XL/XXL — too large"
        if re.search(r'\b(?:size\s+|sz\.?\s+)(?:xl|large)\b', t):
            return True, "workwear size XL/Large — too large"
        if re.search(r'[-–,\s]xl\s*$', t):
            return True, "workwear size XL — too large"
        if re.search(r'\b(?:size\s+|sz\.?\s+)(?:large|medium)\b', t):
            return True, "workwear size Large/Medium — too large"
        if re.search(r'\b(?:size\s+|sz\.?\s+)(?:l|m)\b(?!\s*\d)', t):
            return True, "workwear size L/M — too large"
        if re.search(r'[-–|]\s*(?:size\s+)?(?:large|l)\s*$', t):
            return True, "workwear size L/Large — too large"
        if re.search(r"\bmen'?s\s+(?:size\s+)?(?:large|medium|xl|xxl|l|m)\b(?!ong|arge|\s*\d)", t):
            return True, "workwear Men's L/M/Large/XL — too large"

    # ---- PANTS ---------------------------------------------------------------
    if category == "pants":
        # Explicit waist × inseam format: 30x32, 32x30, 34x32, 36x32 …
        if re.search(r'\b(?:3[0-9]|4[0-9])x\d{2}\b', t):
            return True, "pants waist too large (30+)"
        # W30, W32, W34 … format
        if re.search(r'\bw(?:3[0-9]|4[0-9])\b', t):
            return True, "pants waist too large (W30+)"
        # "Size 36x", "34 x 32" style
        if re.search(r'\bsize\s+(?:3[0-9]|4[0-9])\s*[x×]', t):
            return True, "pants waist too large"

    # ---- TAILORING / OUTERWEAR -----------------------------------------------
    if category == "tailoring_outerwear":
        # Only hard-reject when we can confirm it's a jacket/coat (not a pants or tie)
        is_jacket = bool(re.search(
            r'\b(?:blazer|sport\s*coat|suit(?:\s+jacket)?|overcoat|topcoat|'
            r'peacoat|chesterfield|jacket)\b', t
        ))
        is_trouser = bool(re.search(r'\b(?:trousers?|pants?|chinos?|slacks?)\b', t))
        if is_jacket and not is_trouser:
            # Strip pants waist×inseam dimensions before jacket size checks
            # so "42S 34x28" doesn't count "34" as a valid jacket size.
            tj = re.sub(r'\b\d{2}x\d{2}\b', ' ', t)
            it_luxury = any(b in tj for b in ITALIAN_LUXURY)
            it_reject = _jacket_reject_it_above
            us_reject = _jacket_reject_us_above
            if it_luxury:
                _it_reject_pat = r'\b(?:' + '|'.join(str(n) for n in range(it_reject, 70, 2)) + r')\s*[sr]?\b'
                if re.search(_it_reject_pat, tj):
                    return True, f"Italian jacket IT{it_reject}+ — too large"
                if re.search(r'\b(?:38|40)\s*[sr]?\b', tj):
                    return True, "Italian jacket US 38/40 — too large"
            else:
                _us_accept_sizes = [str(n) for n in range(28, us_reject)]
                _us_accept_pat = r'\b(?:' + '|'.join(_us_accept_sizes) + r')\s*[sr]?\b'
                if not re.search(_us_accept_pat, tj):
                    if re.search(r'\b(?:3[5-9]|4[0-9]|5[0-9])\s*[sr]?\b', tj):
                        return True, f"jacket size {us_reject}+ — too large"
        # Dress shirts pulled in by Italian brand searches (e.g. "Zegna 42"):
        # "42" here is a EUROPEAN COLLAR size (42cm ≈ neck 16.5"), not a jacket.
        # Buyer wears neck 15 (≈38cm); reject collar 41cm+ unless neck 15/15.5 present.
        is_shirt = bool(re.search(r'\b(?:dress\s*shirt|button[\s-]?up|button[\s-]?down|sport\s*shirt|shirt)\b', t)) and not is_jacket
        if is_shirt:
            if re.search(r'\b(?:4[1-9]|5[0-2])\b', t) and not re.search(r'\b(?:15|15\.5|38|39|40)\b', t):
                return True, "dress shirt EU collar 41cm+ — neck too large"
        if is_trouser:
            if re.search(r'\b(?:3[0-9]|4[0-9])x\d{2}\b', t):
                return True, "trouser waist too large (30+)"
            # "Size 36", "Size 34", "Size 32" = waist, reject if not 28
            if re.search(r'\bsize\s+(?:3[0-9]|4[0-9])\b', t) and not re.search(r'\bsize\s+28\b|\b28x\b|\bw28\b', t):
                return True, "trouser waist likely too large (30+)"
            # "36R", "34R" style US waist/length size code on trouser = reject
            if re.search(r'\b(?:3[0-9]|4[0-9])[sr]\b', t) and not re.search(r'\b28[sr]?\b', t):
                return True, "trouser waist size code too large (30R+)"
            # bare number 30-49 anywhere in a trouser title = waist size, reject
            if re.search(r'(?<!\d)(?:3[02-9]|4[0-9])(?!\d)', t) and not re.search(r'(?<!\d)28(?!\d)', t):
                return True, "trouser waist too large (30+)"

    return False, ""


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class Candidate:
    item_id: int
    category: str
    query: str
    search_url: str
    title: str
    url: str
    sgw_category: str
    mens_verification: str
    listing_mode: str
    current_price: Any
    buy_now_price: Any
    shipping_price: Any
    handling_price: Any
    end_time: str
    remaining_time: str
    size: str
    measurements: str
    material: str
    condition_notes: str
    pickup_status: str
    location: str
    image_urls: list[str]
    description_text: str
    live_verified: bool
    match_reasons: list[str] = field(default_factory=list)
    red_flags: list[str] = field(default_factory=list)
    scores: dict[str, int] = field(default_factory=dict)
    total_score: float = 0.0
    recommendation: str = "Skip"


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def clean_text(value: Any) -> str:
    if not value:
        return ""
    if not isinstance(value, str):
        value = json.dumps(value)
    value = re.sub(r"<br\s*/?>", "\n", value, flags=re.I)
    value = re.sub(r"</p>|</li>|</div>", "\n", value, flags=re.I)
    value = re.sub(r"<[^>]+>", " ", value)
    value = html.unescape(value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_image_url(image_server: str, path: str) -> str:
    path = path.replace("\\", "/").strip()
    if not path:
        return ""
    if path.startswith("http"):
        return path
    return image_server.rstrip("/") + "/" + path.lstrip("/")


def search_url(search_text: str, category_id: int) -> str:
    cat = f"&c={category_id}&catIds={category_id}" if category_id else ""
    return (
        f"https://shopgoodwill.com/categories/listing"
        f"?st={quote_plus(search_text)}{cat}&p=1&ps=40&layout=grid"
    )


def default_query(
    search_text: str,
    page: int,
    page_size: int,
    category_id: int = 0,
    category_level: int = 1,
) -> dict[str, Any]:
    return {
        "isSize": False,
        "isWeddingCatagory": "false",
        "isMultipleCategoryIds": False,
        "isFromHeaderMenuTab": False,
        "layout": "grid",
        "searchText": search_text.replace('"', ""),
        "selectedGroup": "",
        "selectedCategoryIds": str(category_id) if category_id else "",
        "selectedSellerIds": "",
        "lowPrice": "0",
        "highPrice": "999999",
        "searchBuyNowOnly": "",
        "searchPickupOnly": "false",
        "searchNoPickupOnly": "false",
        "searchOneCentShippingOnly": "false",
        "searchDescriptions": "false",
        "searchClosedAuctions": "false",
        "closedAuctionEndingDate": datetime.now().strftime("%-m/%-d/%Y"),
        "closedAuctionDaysBack": "7",
        "searchCanadaShipping": "false",
        "searchInternationalShippingOnly": "false",
        "sortColumn": "1",
        "sortDescending": "false",
        "savedSearchId": 0,
        "useBuyerPrefs": "true",
        "searchUSOnlyShipping": "false",
        "categoryLevelNo": str(category_level),
        "categoryLevel": category_level,
        "categoryId": category_id,
        "partNumber": "",
        "catIds": str(category_id) if category_id else "",
        "page": page,
        "pageSize": page_size,
    }


def has_any(text: str, terms: set[str]) -> list[str]:
    lower = text.lower()
    found: set[str] = set()
    for term in terms:
        escaped = re.escape(term.lower()).replace(r"\ ", r"\s+")
        if re.search(rf"(?<![a-z0-9]){escaped}(?![a-z0-9])", lower):
            found.add(term)
    return sorted(found)


# ---------------------------------------------------------------------------
# API calls
# ---------------------------------------------------------------------------

def search_items(
    session: requests.Session,
    query: str,
    page_size: int,
    pages: int,
    category_id: int,
    category_level: int,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for page in range(1, pages + 1):
        resp = session.post(
            f"{API_ROOT}/Search/ItemListing",
            json=default_query(query, page, page_size, category_id, category_level),
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        page_items = data.get("searchResults", {}).get("items", []) or []
        items.extend(page_items)
        if len(page_items) < page_size:
            break
        time.sleep(0.1)
    return items


def get_detail(session: requests.Session, item_id: int) -> dict[str, Any]:
    resp = session.get(
        f"{API_ROOT}/itemDetail/GetItemDetailModelByItemId/{item_id}",
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def fetch_item_html(session: requests.Session, item_id: int) -> str:
    resp = session.get(f"{ITEM_ROOT}/{item_id}", timeout=30)
    resp.raise_for_status()
    return resp.text


# ---------------------------------------------------------------------------
# Field extraction
# ---------------------------------------------------------------------------

def extract_matches(patterns: list[str], text: str) -> str:
    found: list[str] = []
    for pattern in patterns:
        for match in re.findall(pattern, text, flags=re.I):
            value = " ".join(match) if isinstance(match, tuple) else match
            value = re.sub(r"\s+", " ", value).strip(" :;-")
            if value and value.lower() not in {x.lower() for x in found}:
                found.append(value)
    return "; ".join(found[:10])


def extract_fields(detail: dict[str, Any]) -> tuple[str, str, str, str]:
    text = clean_text(detail.get("description"))
    combined = text + " " + (detail.get("title") or "")
    size = extract_matches([
        r"\b(?:US Size|Size|Tagged Size|Neck|Waist|W)\s*:?\s*([0-9]{1,2}(?:\.[05])?(?:\s?[A-Z])?)",
        r"\b(36S|36R|36|Italian 44|EU 44|44R|Small|size small|15\.5|15|28x30|28x32|W28|28 waist)\b",
    ], combined)
    measurements = extract_matches([
        r"\b(pit(?: to pit)?|p2p|chest|shoulder|sleeve|length|waist|inseam|rise)\s*:?\s*([0-9]{1,2}(?:\.[0-9])?(?:\s?(?:in|inch|\"))?)",
        r"\b([0-9]{1,2}(?:\.[0-9])?)\s?(?:in|inch|\")\s+(pit|chest|shoulder|sleeve|length|waist|inseam)",
    ], text)
    material = extract_matches([
        r"\b(?:Material|Upper Material|Fabric|Shell|Composition)\s*:?\s*([A-Za-z /%.-]{3,80})",
        r"\b(cashmere|camel hair|camelhair|harris tweed|tweed|wool|linen|silk|alpaca|mohair|merino|corduroy|shell cordovan|calfskin|leather|cotton|moleskin)\b",
    ], text)
    condition = extract_matches([
        r"\b(?:Condition)\s*:?\s*([A-Za-z /%.,'-]{3,120})",
        r"\b(stain(?:s|ing)?|hole(?:s)?|moth|odor(?:s)?|yellowing|wear|scuff(?:s|ing)?|tear(?:s)?|damage|missing button(?:s)?|pre-owned|used)\b",
    ], text)
    return size, measurements, material, condition


# ---------------------------------------------------------------------------
# Men's verification
# ---------------------------------------------------------------------------

def mens_status(
    category: str,
    detail: dict[str, Any],
    size: str,
    measurements: str,
) -> tuple[str, list[str]]:
    if category == "blankets":
        return "Men's", ["blanket/throw is not gendered apparel"]

    cat_text = (
        clean_text(detail.get("categoryParentList"))
        + " "
        + clean_text(detail.get("categoryBreadCrumbs"))
    )
    text = " ".join([
        detail.get("title") or "",
        clean_text(detail.get("description")),
        cat_text,
        size,
        measurements,
    ]).lower()

    evidence: list[str] = []
    likely_women = False
    uncertain = False

    if re.search(r"\bmen'?s\b|\bmens\b", text):
        evidence.append("men's term in title/category/description")
    if "men's clothing" in text or "shoes men's" in text or "men's shoes" in text:
        evidence.append("ShopGoodwill men's category confirmed")
    if re.search(r"\bwomen'?s\b|\bwomens\b|\blad(y|ies)\b|\bgirl\b|\byouth\b", text):
        evidence.append("women's/youth term found")
        likely_women = True
    if re.search(r"\b(pump|pumps|heel|heels|stiletto|skirt|gown|blouse|petite)\b", text):
        evidence.append("women-coded garment term found")
        likely_women = True
    if category == "shoes" and re.search(r"\b8\s?(aaaa|aaa|aa|b)\b|\bsize 38\b", text):
        evidence.append("size format commonly women's")
        uncertain = True
    if category in {"shirts", "tailoring_outerwear", "pants", "workwear", "knitwear"} \
            and "women" not in text \
            and not re.search(r"\bmen'?s\b|\bmens\b|men's clothing", text):
        evidence.append("men's wording not explicit")
        uncertain = True
    if category == "pants" \
            and re.search(r"\binseam\s*:?\s*28\b", text) \
            and not re.search(r"\b(w28|28 waist|waist\s*:?\s*28|28x\d{2})\b", text):
        evidence.append("28 may be inseam not waist")
        uncertain = True

    if likely_women:
        return "Likely Women's", evidence
    if uncertain:
        return "Uncertain", evidence
    if evidence:
        return "Men's", evidence
    return "Likely Men's", ["no women's signals; fit/category signals compatible with menswear"]


def listing_mode(detail: dict[str, Any]) -> str:
    if detail.get("isStock"):
        return "Stock / Buy Now"
    if detail.get("buyNowPrice") or detail.get("discountedBuyNowPrice"):
        return "Auction with Buy Now"
    if detail.get("isAuction"):
        return "Auction"
    return "not verified"


# ---------------------------------------------------------------------------
# Scoring / assessment
# ---------------------------------------------------------------------------

def assess(
    category: str,
    title: str,
    query: str,
    detail: dict[str, Any],
    size: str,
    measurements: str,
    material: str,
    condition: str,
    verified_mens_status: str,
) -> tuple[list[str], list[str], dict[str, int], str]:
    text = " ".join([
        title,
        clean_text(detail.get("description")),
        clean_text(detail.get("categoryParentList")),
        clean_text(detail.get("categoryBreadCrumbs")),
        size, measurements, material, condition,
    ]).lower()

    reasons: list[str] = []
    red_flags: list[str] = []
    brands = has_any(text, QUALITY_BRANDS)
    fabrics = has_any(text, FABRIC_TERMS)
    mall = has_any(text, MALL_BRANDS)

    if brands:
        reasons.append("maker match: " + ", ".join(brands[:5]))
    if fabrics:
        reasons.append("fabric match: " + ", ".join(fabrics[:5]))

    # --- fit ---
    fit = 3
    size_text = " ".join([title, size, measurements]).lower()
    shoe_ok = bool(_shoe_accept_re.search(size_text))
    # Strip waist×inseam pants dimensions (e.g. "34x28", "34x30") before checking
    # jacket sizes — otherwise pants dimensions masquerade as jacket size matches.
    jacket_text = re.sub(r'\b\d{2}x\d{2}\b', ' ', text)
    if category == "shoes" and shoe_ok:
        fit = 8
        reasons.append(f"shoe size in target range ({'/'.join(BUYER_PROFILE['shoe_sizes'])})")
    elif category in {"tailoring_outerwear", "workwear"} and _jacket_accept_re.search(jacket_text):
        fit = 8
        reasons.append(
            "tailoring size in target range ("
            + "/".join(BUYER_PROFILE["jacket_sizes_us"])
            + " / IT" + "/".join(BUYER_PROFILE["jacket_sizes_it"]) + ")"
        )
    elif category == "pants" and re.search(
        rf'\b({re.escape(_pants_waist)} waist|w{re.escape(_pants_waist)}|'
        rf'{re.escape(_pants_waist)}x\d\d|\b{re.escape(_pants_waist)}\b)\b', text
    ):
        fit = 8
        reasons.append(f"waist in target range ({_pants_waist}\")")
    elif category in {"shirts", "knitwear"} and re.search(
        rf'\b({re.escape(_shirt_neck)}|{re.escape(_body_size)})\b', text
    ):
        # Neck exception size only accepted for exceptional makers
        if re.search(rf'\b{re.escape(_shirt_neck_exception)}\b', text):
            if any(m in text for m in EXCEPTIONAL_MAKERS_15_5):
                fit = 8
                reasons.append(f"exceptional maker at {_shirt_neck_exception} neck — valid exception")
            else:
                fit = 2
                red_flags.append("15.5 neck — not an exceptional maker (size out of primary range)")
        else:
            fit = 8
            reasons.append("shirt/knit size in target range (15 neck)")
    elif category == "blankets":
        fit = 7
        reasons.append("blanket sizing is not body-fit dependent")

    # --- price ---
    current_price = detail.get("currentPrice") or 0
    buy_now = detail.get("discountedBuyNowPrice") or detail.get("buyNowPrice") or 0
    visible_price = buy_now if buy_now else current_price
    if visible_price and visible_price <= 30:
        price_value = 8
    elif visible_price and visible_price <= 75:
        price_value = 6
    elif visible_price and visible_price > 150:
        price_value = 3
    else:
        price_value = 5

    quality = min(10, 3 + min(len(brands), 3) * 2 + min(len(fabrics), 3))
    rarity = min(10, 3 + len(brands) + len(fabrics))
    taste = min(10, 4 + len(fabrics) + (2 if brands else 0))
    risk = 7

    # --- penalise ---
    if mall:
        red_flags.append("low-upside brand: " + ", ".join(mall))
        quality = min(quality, 4)
    if verified_mens_status in {"Uncertain", "Likely Women's"}:
        red_flags.append(f"men's verification: {verified_mens_status}")
        fit = min(fit, 3 if verified_mens_status == "Uncertain" else 1)
    if re.search(r"\b(women|womens|women's|ladies|lady|girl|youth)\b", text) and category != "blankets":
        red_flags.append("possible women's listing")
        fit = min(fit, 2)
    if category == "shoes":
        if re.search(r"\b(shoe tree|wallet|handbag|purse|heel|pump)\b", title.lower()):
            red_flags.append("non-shoe result in shoe search")
            fit = min(fit, 2)
        if not re.search(r"\b(shoe|shoes|loafer|oxford|derby|boot|brogue|monk|moc|sneaker|dress)\b", text):
            red_flags.append("shoe type not confirmed in detail")
            fit = min(fit, 2)
        if not shoe_ok:
            red_flags.append("target shoe size not confirmed")
            fit = min(fit, 3)
        if re.search(r"(?<![\d.])(?:8\.5|9|10|11|12)(?:\s?[a-z]{0,2})?(?![\d.])", size_text) and not shoe_ok:
            red_flags.append("shoe size appears outside target")
            fit = min(fit, 2)
    if detail.get("pickupOnly") or detail.get("storePickupOnly"):
        red_flags.append("pickup-only")
        risk -= 3
    if re.search(r"\b(stain|hole|moth|odor|yellowing|damage|tear)\b", text):
        red_flags.append("condition issue mentioned")
        risk -= 2
    if not size and category != "blankets":
        red_flags.append("size not confirmed in detail text")
        fit = min(fit, 4)
    if not material and category in {"tailoring_outerwear", "workwear", "pants", "shirts", "blankets", "knitwear"}:
        red_flags.append("fabric/material not confirmed")

    maker_quality = min(10, 2 + min(len(brands), 4) * 2)
    fabric_quality = min(10, 2 + min(len(fabrics), 5) * 2)
    scores = {
        "fit_likelihood": max(1, min(10, fit)),
        "fabric_quality": max(1, min(10, fabric_quality)),
        "maker_quality": max(1, min(10, maker_quality)),
        "condition": max(1, min(10, risk)),
        "value": max(1, min(10, price_value)),
        "rarity_upside": max(1, min(10, rarity)),
        "quiet_taste_fit": max(1, min(10, taste)),
    }

    avg = sum(scores.values()) / len(scores)
    severe = any(
        marker in flag
        for flag in red_flags
        for marker in ["women", "youth", "non-shoe", "Uncertain", "pickup-only"]
    )
    if severe:
        recommendation = "Skip"
    elif avg >= 7.3:
        recommendation = "Buy"
    elif avg >= 5.8:
        recommendation = "Watch"
    elif avg >= 4.8:
        recommendation = "Need measurements"
    else:
        recommendation = "Skip"

    if not reasons:
        reasons.append("keyword match only; detail page has limited supporting evidence")
    return reasons, red_flags, scores, recommendation


# ---------------------------------------------------------------------------
# Candidate builder
# ---------------------------------------------------------------------------

def build_candidate(
    category: str,
    query: str,
    item: dict[str, Any],
    detail: dict[str, Any],
) -> Candidate:
    item_id = int(detail.get("itemId") or item.get("itemId"))
    title = detail.get("title") or item.get("title") or ""
    size, measurements, material, condition = extract_fields(detail)
    verified_mens_status, mens_evidence = mens_status(category, detail, size, measurements)
    image_server = detail.get("imageServer") or ""
    image_paths = (detail.get("imageUrlString") or item.get("imageURL") or "").split(";")
    image_urls = [
        normalize_image_url(image_server, p)
        for p in image_paths
        if normalize_image_url(image_server, p)
    ]
    pickup_bits = []
    if detail.get("pickupOnly") or detail.get("storePickupOnly"):
        pickup_bits.append("pickup only")
    elif detail.get("sellerAllowPickup"):
        pickup_bits.append("pickup available")
    if detail.get("noInternationalShippingMessage"):
        pickup_bits.append(detail["noInternationalShippingMessage"])
    location = ", ".join(
        x for x in [detail.get("pickupCity"), detail.get("pickupState"), detail.get("pickupZip")] if x
    )
    reasons, red_flags, scores, recommendation = assess(
        category, title, query, detail, size, measurements, material, condition, verified_mens_status
    )
    reasons.extend(f"men's check: {e}" for e in mens_evidence[:2])
    total_score = round(sum(scores.values()) / len(scores), 2)
    live_verified = (
        not bool(detail.get("isItemEndTimeExpire"))
        or bool(detail.get("buyNowPrice") or detail.get("discountedBuyNowPrice"))
    )
    return Candidate(
        item_id=item_id,
        category=category,
        query=query,
        search_url=search_url(query, SEARCH_GROUPS[category]["category_id"]),
        title=title,
        url=f"{ITEM_ROOT}/{item_id}",
        sgw_category=(
            clean_text(detail.get("categoryParentList"))
            or clean_text(detail.get("categoryBreadCrumbs"))
            or item.get("catFullName")
            or "not verified"
        ),
        mens_verification=verified_mens_status,
        listing_mode=listing_mode(detail),
        current_price=detail.get("currentPrice"),
        buy_now_price=detail.get("discountedBuyNowPrice") or detail.get("buyNowPrice"),
        shipping_price=detail.get("shippingPrice") or "not verified",
        handling_price=detail.get("handlingPrice") or "not verified",
        end_time=detail.get("endTime") or item.get("endTime") or "",
        remaining_time=detail.get("remainingTime") or item.get("remainingTime") or "",
        size=size or "not verified",
        measurements=measurements or "not verified",
        material=material or "not verified",
        condition_notes=condition or "not verified",
        pickup_status="; ".join(pickup_bits) or "not verified",
        location=location or "not verified",
        image_urls=image_urls,
        description_text=clean_text(detail.get("description")),
        live_verified=live_verified,
        match_reasons=reasons,
        red_flags=red_flags,
        scores=scores,
        total_score=total_score,
        recommendation=recommendation if live_verified else "Skip",
    )


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def candidate_row(c: Candidate) -> dict[str, Any]:
    row = c.__dict__.copy()
    row["image_urls"] = " | ".join(c.image_urls)
    row["match_reasons"] = " | ".join(c.match_reasons)
    row["red_flags"] = " | ".join(c.red_flags)
    for k, v in c.scores.items():
        row[f"score_{k}"] = v
    row.pop("scores")
    return row


def write_report(
    candidates: list[Candidate],
    output_dir: Path,
    search_log: list[dict[str, Any]],
) -> None:
    top = sorted(
        [c for c in candidates if c.live_verified and c.recommendation in {"Buy", "Watch"} and c.mens_verification in {"Men's", "Likely Men's"}],
        key=lambda c: c.total_score, reverse=True,
    )[:20]
    maybe = sorted(
        [c for c in candidates if c.live_verified and c.recommendation == "Need measurements" and c.mens_verification in {"Men's", "Likely Men's"}],
        key=lambda c: c.total_score, reverse=True,
    )[:25]
    rejected = sorted(
        [c for c in candidates if c not in top and c not in maybe],
        key=lambda c: c.total_score, reverse=True,
    )[:40]

    lines = [
        "# ShopGoodwill Menswear Hunt",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"Verified item-detail pages: {len(candidates)}",
        f"Atomic search strings run: {len(search_log)}",
        f"Strings with live results: {sum(1 for e in search_log if e.get('count', 0) > 0)}",
        "",
        "## Top Picks",
    ]
    if not top:
        lines.append("No verified Buy/Watch picks met the scoring threshold.")
    for i, c in enumerate(top, 1):
        lines += [
            f"### {i}. [{c.title}]({c.url})",
            f"- **Item ID:** {c.item_id}",
            f"- **Search string:** `{c.query}`",
            f"- **Category:** {c.category}; {c.sgw_category}",
            f"- **Men's verification:** {c.mens_verification}",
            f"- **Score:** {c.total_score}/10 — **{c.recommendation}**",
            f"- **Listing:** {c.listing_mode}",
            f"- **Price:** current ${c.current_price} / Buy Now ${c.buy_now_price or 0} / shipping {c.shipping_price} / handling {c.handling_price}",
            f"- **Ends:** {c.end_time} PT — remaining {c.remaining_time or 'not verified'}",
            f"- **Size:** {c.size}",
            f"- **Measurements:** {c.measurements}",
            f"- **Material:** {c.material}",
            f"- **Condition:** {c.condition_notes}",
            f"- **Location:** {c.location}; {c.pickup_status}",
            f"- **Why it matches:** {'; '.join(c.match_reasons)}",
            f"- **Red flags:** {'; '.join(c.red_flags) if c.red_flags else 'none'}",
            f"- **Images:** {'; '.join(c.image_urls[:3]) if c.image_urls else 'none'}",
            "",
        ]

    lines.append("## Maybe / Need Measurements")
    for c in maybe:
        lines.append(
            f"- [{c.title}]({c.url}) — {c.total_score}/10, {c.mens_verification}, "
            f"`{c.query}`, size: {c.size}, material: {c.material}; "
            f"flags: {'; '.join(c.red_flags) or 'none'}"
        )

    lines += ["", "## Rejected But Notable"]
    for c in rejected:
        lines.append(
            f"- [{c.title}]({c.url}) — {c.total_score}/10, {c.mens_verification}, "
            f"`{c.query}`; {'; '.join(c.red_flags) or 'lower score'}"
        )

    (output_dir / "report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


# Categories that get their own separate CSV (non-clothing / home goods)
NON_CLOTHING_CATEGORIES = {"blankets"}


def write_csv(rows: list[dict[str, Any]], path: Path) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def save_outputs(
    candidates: list[Candidate],
    output_dir: Path,
    search_log: list[dict[str, Any]],
) -> None:
    all_rows = [candidate_row(c) for c in candidates]
    (output_dir / "candidates.json").write_text(json.dumps(all_rows, indent=2), encoding="utf-8")

    clothing_rows = [r for r in all_rows if r["category"] not in NON_CLOTHING_CATEGORIES]
    write_csv(clothing_rows, output_dir / "candidates.csv")

    # Separate CSV per non-clothing category
    for cat in NON_CLOTHING_CATEGORIES:
        cat_rows = [r for r in all_rows if r["category"] == cat]
        if cat_rows:
            write_csv(cat_rows, output_dir / f"{cat}.csv")
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            colors = {"Buy": "C6EFCE", "Watch": "FFEB9C", "Need measurements": "DDEBF7"}

            def make_sheet(wb_obj: Any, title: str, sheet_rows: list[dict[str, Any]]) -> None:
                ws = wb_obj.create_sheet(title=title)
                if not sheet_rows:
                    return
                headers = list(sheet_rows[0].keys())
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = PatternFill("solid", fgColor="1F4E79")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(wrap_text=False)
                for row_idx, row in enumerate(sheet_rows, 2):
                    fill_color = colors.get(row.get("recommendation", ""))
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col, value=row[h])
                        if fill_color:
                            cell.fill = PatternFill("solid", fgColor=fill_color)
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

            # Clothing workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            make_sheet(wb, "Clothing", clothing_rows)
            wb.save(output_dir / "candidates.xlsx")

            # Non-clothing workbook (one sheet per category)
            non_clothing = [r for r in all_rows if r["category"] in NON_CLOTHING_CATEGORIES]
            if non_clothing:
                wb2 = openpyxl.Workbook()
                wb2.remove(wb2.active)
                for cat in NON_CLOTHING_CATEGORIES:
                    cat_rows = [r for r in non_clothing if r["category"] == cat]
                    if cat_rows:
                        make_sheet(wb2, cat.capitalize(), cat_rows)
                wb2.save(output_dir / "blankets_home.xlsx")
        except ImportError:
            pass  # openpyxl not installed; CSV is the fallback
    write_report(candidates, output_dir, search_log)


def save_evidence(
    session: requests.Session,
    candidates: list[Candidate],
    output_dir: Path,
    limit: int,
) -> None:
    evidence_dir = output_dir / "evidence"
    evidence_dir.mkdir(parents=True, exist_ok=True)
    top = [
        c for c in sorted(candidates, key=lambda c: c.total_score, reverse=True)
        if c.recommendation in {"Buy", "Watch"} and c.live_verified
    ][:limit]
    by_id = {c.item_id: c for c in candidates}
    for c in top:
        try:
            html_text = fetch_item_html(session, c.item_id)
            (evidence_dir / f"{c.item_id}.html").write_text(html_text, encoding="utf-8")
        except Exception:
            pass  # item may have ended or returned an error; skip silently
        (evidence_dir / f"{c.item_id}.txt").write_text(
            json.dumps(candidate_row(by_id[c.item_id]), indent=2), encoding="utf-8"
        )
        time.sleep(0.2)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="ShopGoodwill menswear hunter.")
    parser.add_argument("--output", default="runs/latest")
    parser.add_argument("--per-query", type=int, default=10)
    parser.add_argument("--pages-per-query", type=int, default=1)
    parser.add_argument("--max-detail", type=int, default=2000)
    parser.add_argument("--evidence-limit", type=int, default=20)
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    output_dir = Path(args.output)
    if not output_dir.is_absolute():
        output_dir = script_dir / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    seen: set[int] = set()
    candidates: list[Candidate] = []
    search_log: list[dict[str, Any]] = []
    total_strings = sum(len(g["strings"]) for g in SEARCH_GROUPS.values())
    done = 0

    for category, group in SEARCH_GROUPS.items():
        cat_id = int(group["category_id"])
        cat_level = int(group["category_level"])
        for query in group["strings"]:
            done += 1
            try:
                items = search_items(session, query, args.per_query, args.pages_per_query, cat_id, cat_level)
                search_log.append({
                    "category": category,
                    "query": query,
                    "search_url": search_url(query, cat_id),
                    "count": len(items),
                })
                if items:
                    print(f"[{done}/{total_strings}] {category}: '{query}' → {len(items)} results")
            except Exception as exc:
                search_log.append({"category": category, "query": query, "error": str(exc)})
                print(f"[{done}/{total_strings}] {category}: '{query}' → ERROR: {exc}")
                continue
            for item in items:
                item_id = int(item.get("itemId") or 0)
                if not item_id or item_id in seen:
                    continue
                seen.add(item_id)
                if len(candidates) >= args.max_detail:
                    break
                title = item.get("title") or ""
                skip, skip_reason = pre_fetch_reject(category, title)
                if skip:
                    search_log.append({
                        "category": category, "query": query,
                        "item_id": item_id, "pre_fetch_skip": skip_reason, "title": title,
                    })
                    continue
                try:
                    detail = get_detail(session, item_id)
                    candidate = build_candidate(category, query, item, detail)
                    candidates.append(candidate)
                except Exception as exc:
                    search_log.append({
                        "category": category, "query": query,
                        "item_id": item_id, "detail_error": str(exc),
                    })
                time.sleep(0.15)
            if len(candidates) >= args.max_detail:
                break
        if len(candidates) >= args.max_detail:
            break

    candidates.sort(key=lambda c: c.total_score, reverse=True)
    save_outputs(candidates, output_dir, search_log)
    save_evidence(session, candidates, output_dir, args.evidence_limit)
    (output_dir / "search_log.json").write_text(json.dumps(search_log, indent=2), encoding="utf-8")
    print(f"\nDone. {len(candidates)} candidates → {output_dir}")


if __name__ == "__main__":
    main()
