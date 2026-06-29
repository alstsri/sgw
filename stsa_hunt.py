#!/usr/bin/env python3
"""ShopTheSalvationArmy.com menswear hunter — HTML search, browser headers."""
import argparse
import csv
import json
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus

import requests

BASE_URL = "https://www.shopthesalvationarmy.com"
ITEM_ROOT = f"{BASE_URL}/Listing/Details"

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
}

# ---------------------------------------------------------------------------
# BUYER PROFILE — edit this block to configure for a different buyer.
# ---------------------------------------------------------------------------
BUYER_PROFILE = {
    "shoe_sizes": ["8", "7.5"],
    "shoe_widths": ["", "D", "M", "E"],
    "shoe_reject_above": 8.5,
    "jacket_sizes_us": ["34S", "34R", "34", "32R", "32"],
    "jacket_sizes_it": ["42", "44"],
    "jacket_reject_us_above": 35,
    "jacket_reject_it_above": 46,
    "shirt_neck": "15",
    "shirt_neck_exception": "15.5",
    "pants_waist": "28",
    "body_size": "small",
}

# ---------------------------------------------------------------------------
# Brand / fabric intelligence (same as hunt.py)
# ---------------------------------------------------------------------------

QUALITY_BRANDS = {
    "alden", "allen edmonds", "red wing", "white's", "wesco", "viberg",
    "carmina", "carlos santos", "meermin", "santoni", "testoni", "bontoni",
    "magnanni", "tod's", "bally", "ferragamo", "stefano bemer", "silvano lattanzi",
    "esquivel", "quoddy", "vass", "sanders", "visvim",
    "crockett", "tricker", "edward green", "john lobb", "gaziano", "church",
    "cheaney", "grenson", "loake", "alfred sargent", "george cleverley",
    "weston", "paraboot", "berluti", "heschung", "corthay", "aubercy",
    "loro piana", "brunello", "zegna", "ermenegildo", "canali", "corneliani",
    "brioni", "kiton", "isaia", "belvest", "sartorio", "boglioli", "caruso",
    "ring jacket", "attolini", "cesare attolini",
    "anderson & sheppard", "henry poole", "huntsman", "gieves", "crombie",
    "hackett", "cordings", "daks", "aquascutum", "barbour", "belstaff",
    "private white", "mackintosh", "grenfell",
    "arnys", "cifonelli", "smalto", "husbands paris", "officine generale",
    "lemaire", "hermes", "hermès",
    "oxxford", "hickey freeman", "golden fleece", "j. press", "j press",
    "southwick", "samuelsohn", "paul stuart", "ralph lauren", "rrl", "pendleton",
    "harris tweed", "schott", "filson",
    "incotex", "zanella", "pt01", "pt torino", "berwich", "rota",
    "turnbull", "asser", "charvet", "eton", "barba", "finamore", "borrelli",
    "fray", "lorenzini", "gitman", "kamakura", "hilditch", "harvie",
    "saint james", "drumohr", "inis meain", "william lockie", "johnstons",
    "john smedley", "ballantyne", "pringle", "malo",
    "vetra", "le laboureur",
    "faribault", "woolrich", "pendleton",
}

FABRIC_TERMS = {
    "cashmere", "camel hair", "wool", "harris tweed", "tweed", "flannel",
    "alpaca", "mohair", "linen", "silk", "shell cordovan", "cordovan",
    "calfskin", "goodyear", "made in england", "made in italy", "made in usa",
    "made in scotland", "made in ireland", "made in japan", "made in france",
    "merino", "sea island", "donegal",
}

REJECT_BRANDS = {
    "clarks", "bass", "florsheim", "nunn bush", "rockport", "skechers",
    "steve madden", "aldo", "bostonian", "stafford", "lands' end",
    "van heusen", "arrow", "alfani", "dockers", "izod", "express",
    "wrangler", "topman",
}

MALL_BRANDS = {
    "zara", "h&m", "shein", "forever 21", "cole haan", "banana republic",
}

# ---------------------------------------------------------------------------
# Search strings — same buyer profile as hunt.py
# ---------------------------------------------------------------------------

SEARCH_STRINGS = [
    # Shoes — American/Italian
    "Allen Edmonds 8", "Allen Edmonds 7.5",
    "Alden 8", "Alden 7.5",
    "Red Wing 8", "Red Wing 7.5",
    "Ferragamo 8", "Ferragamo 7.5", "Salvatore Ferragamo 8",
    "Santoni 8", "Santoni 7.5",
    "Magnanni 8", "Magnanni 7.5",
    "Tod's 8", "Tod's 7.5",
    "Carmina 8", "Carmina 7.5",
    "Quoddy 8", "Quoddy 7.5",
    "Esquivel 8", "Esquivel 7.5",
    "Viberg 8", "Viberg 7.5",
    # Shoes — English
    "Crockett Jones 8", "Crockett & Jones 8",
    "Edward Green 8", "John Lobb 8",
    "Church's 8", "Tricker's 8",
    "Grenson 8", "Loake 1880 8",
    # Shoes — French
    "JM Weston 8", "Paraboot 8", "Berluti 8",
    # Tailoring — Italian (IT42/IT44)
    "Canali 42", "Canali 44",
    "Corneliani 42", "Corneliani 44",
    "Zegna 42", "Zegna 44", "Ermenegildo Zegna 42",
    "Brioni 42", "Brioni 44",
    "Kiton 42", "Kiton 44",
    "Isaia 42", "Boglioli 42", "Sartorio 42", "Caruso 42",
    "Loro Piana 42", "Brunello Cucinelli 42",
    # Tailoring — American/English (US 34S/34R)
    "Hickey Freeman 34", "Oxxford 34",
    "Brooks Brothers Golden Fleece 34",
    "J Press 34", "Southwick 34",
    "Paul Stuart 34", "Ralph Lauren Purple Label 34",
    "Hackett 34", "Crombie 34",
    "camel hair coat 34", "cashmere coat 34",
    "overcoat 34", "topcoat 34", "trench coat 34",
    "shearling jacket 34", "sheepskin jacket 34",
    "tweed sport coat 34", "houndstooth 34",
    # Workwear
    "Filson Mackinaw", "Vetra chore coat", "Le Laboureur",
    "French workwear", "Pendleton wool shirt",
    # Knitwear
    "Drumohr", "Inis Meain", "John Smedley small",
    "Ballantyne cashmere", "Pringle Scotland",
    "Loro Piana sweater", "Brunello Cucinelli sweater",
    # Shirts — neck 15
    "Charvet shirt 15", "Hermes shirt 15",
    "Turnbull Asser shirt 15", "Brioni shirt 15",
    "Kiton shirt 15", "Finamore shirt 15",
    "Eton shirt 15", "Gitman shirt 15",
    # Accessories / ties
    "Drake's tie", "Hermes tie", "Brioni tie",
    "E. Marinella tie", "Kiton tie",
    # Blankets
    "Pendleton wool blanket", "Faribault wool blanket",
    "Hudson Bay point blanket",
    "merino wool throw", "cashmere throw",
]

# ---------------------------------------------------------------------------
# HTML parsing helpers
# ---------------------------------------------------------------------------

def _strip_tags(text: str) -> str:
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _find(pattern: str, html: str, group: int = 1, flags: int = re.I) -> str:
    m = re.search(pattern, html, flags)
    if not m:
        return ""
    try:
        return m.group(group).strip()
    except IndexError:
        return m.group(0).strip()


def parse_gallery_page(html: str) -> list[dict[str, Any]]:
    """Extract listing stubs from a search results page."""
    items: list[dict[str, Any]] = []
    # Each gallery unit: <div ... data-listingid="ID" ...>...(block)...</div></div></div>
    # Split on galleryUnit boundaries
    units = re.split(r'(?=<div[^>]+class="[^"]*galleryUnit[^"]*")', html)
    for unit in units:
        lid = _find(r'data-listingid="(\d+)"', unit)
        if not lid:
            continue
        # Title from img alt (most reliable)
        title_raw = _find(r'img[^>]+alt="([^"]{3,150})"', unit)
        if not title_raw:
            # Fallback: text after <br /> in gallery title
            title_raw = _find(r'<br\s*/?>\s*\n?\s*([^\n<]{5,200})', unit)
        title = unescape(title_raw).strip()
        price = _find(r'NumberPart">([^<]+)', unit)
        bids = _find(r'awe-rt-AcceptedListingActionCount[^>]*>(\d+)', unit)
        href = _find(r'href="(/Listing/Details/[^"]+)"', unit)
        status_paused = bool(re.search(r'ShowStatusPaused(?!\s*awe-hidden)', unit))
        # Only flag closed when it appears as visible content, not in data-* attributes
        status_closed = bool(re.search(r'(?:awe-rt-ShowStatusEnded(?!\s*awe-hidden)|class="[^"]*auction-closed[^"]*")', unit, re.I))
        img = _find(r'shopsalvationarmyblob\.blob\.core\.windows\.net/assets/media/[^\'"]+', unit)
        items.append({
            "item_id": int(lid),
            "title": title,
            "url": BASE_URL + href if href else f"{ITEM_ROOT}/{lid}",
            "current_price": float(price.replace(",", "")) if price else None,
            "num_bids": int(bids) if bids else 0,
            "image_url": f"https://{img}" if img else "",
            "status_paused": status_paused,
            "status_closed": status_closed,
        })
    return items


def parse_detail_page(html: str, item_id: int) -> dict[str, Any]:
    """Extract full detail from a listing detail page."""
    # Title from og:title (cleanest — no nav buttons)
    title = _find(r'og:title[^>]*content="([^"]+)"', html, flags=re.I)
    if not title:
        # Fallback: h1, stripping child tags
        h1 = re.search(r'<h1[^>]*>(.*?)</h1>', html, re.S)
        title = _strip_tags(h1.group(1)).split('\n')[0].strip() if h1 else ""

    current_price_str = _find(r"awe-rt-CurrentPrice[^>]*>[^$]*\$<span[^>]*>([\d,\.]+)", html)
    current_price = float(current_price_str.replace(",", "")) if current_price_str else None
    buy_now_str = _find(r'buyNowPriceForJS\s*=\s*"([^"]*)"', html)
    buy_now = float(buy_now_str) if buy_now_str and buy_now_str not in ("", "0", "0.00") else None
    # End time is JS-loaded via SignalR — not available in static HTML
    end_time = ""
    num_bids = _find(r'awe-rt-AcceptedListingActionCount[^>]*>(\d+)', html)
    shipping = _find(r'[Ss]hipping[^:]*:[^$\n]*\$([\d\.]+)', html)
    is_closed = bool(re.search(r'(?:awe-rt-ShowStatusEnded(?!\s*awe-hidden)|This auction has ended)', html, re.I))

    # Structured description block: <ul><li><strong>Field:</strong> value</li>...</ul>
    desc_block = re.search(r'class="[^"]*detail__sectionBody description[^"]*"[^>]*>(.*?)</div>', html, re.S | re.I)
    if desc_block:
        desc_html = desc_block.group(1)
        # Extract key-value pairs
        fields = re.findall(r'<strong>([^<]+)</strong>\s*:?\s*([^<\n]{2,120})', desc_html)
        description = "; ".join(f"{k.strip()}: {_strip_tags(v).strip()}" for k, v in fields if v.strip())
        # Extract size from structured data
        size_field = next((v for k, v in fields if "size" in k.lower()), "")
        condition_field = next((v for k, v in fields if "condition" in k.lower()), "")
        brand_field = next((v for k, v in fields if "brand" in k.lower()), "")
    else:
        description = ""
        size_field = ""
        condition_field = _find(r'[Cc]ondition[^:]*:</[a-z]+>\s*([^<\n]{3,60})', html)
        brand_field = ""

    # Images
    images = list(dict.fromkeys(
        f"https://{i}" for i in re.findall(
            r'shopsalvationarmyblob\.blob\.core\.windows\.net/assets/media/[^\'"&\s]+', html
        ) if "_fullsize" in i or "_largesize" in i
    ))

    return {
        "item_id": item_id,
        "title": title,
        "url": f"{ITEM_ROOT}/{item_id}",
        "current_price": current_price,
        "buy_now_price": buy_now,
        "end_time": end_time,
        "num_bids": int(num_bids) if num_bids else 0,
        "shipping": float(shipping) if shipping else None,
        "condition": condition_field,
        "size": _strip_tags(size_field).strip(),
        "brand": _strip_tags(brand_field).strip(),
        "description": description[:600],
        "image_urls": images[:5],
        "live_verified": not is_closed,
    }

# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def search_page(session: requests.Session, query: str, page: int) -> list[dict[str, Any]]:
    url = f"{BASE_URL}/Search?query={quote_plus(query)}&page={page}"
    resp = session.get(url, timeout=20)
    resp.raise_for_status()
    return parse_gallery_page(resp.text)


def fetch_detail(session: requests.Session, item_id: int) -> dict[str, Any]:
    resp = session.get(f"{ITEM_ROOT}/{item_id}", timeout=20)
    resp.raise_for_status()
    return parse_detail_page(resp.text, item_id)

# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

def has_any(text: str, terms: set[str]) -> list[str]:
    lower = text.lower()
    return sorted(t for t in terms if re.search(rf"(?<![a-z0-9]){re.escape(t.lower())}(?![a-z0-9])", lower))


def score_candidate(detail: dict[str, Any], query: str) -> dict[str, Any]:
    text = " ".join([
        detail.get("title") or "",
        detail.get("description") or "",
        detail.get("condition") or "",
    ]).lower()

    brands = has_any(text, QUALITY_BRANDS)
    fabrics = has_any(text, FABRIC_TERMS)
    mall = has_any(text, MALL_BRANDS)
    reject = has_any(text, REJECT_BRANDS)

    reasons: list[str] = []
    red_flags: list[str] = []

    if brands:
        reasons.append("maker: " + ", ".join(brands[:4]))
    if fabrics:
        reasons.append("fabric: " + ", ".join(fabrics[:4]))
    if reject:
        red_flags.append("reject brand: " + ", ".join(reject))
    if mall:
        red_flags.append("low-upside brand: " + ", ".join(mall))
    if detail.get("status_paused"):
        red_flags.append("auction paused")

    # Fit score (coarse — based on query and title text)
    fit = 3
    sz_text = (detail.get("title") or "").lower()
    shoe_sizes = BUYER_PROFILE["shoe_sizes"]
    if any(s in sz_text for s in shoe_sizes):
        fit = 8
        reasons.append(f"shoe size in target ({'/'.join(shoe_sizes)})")
    elif any(s.lower() in sz_text for s in BUYER_PROFILE["jacket_sizes_us"] + BUYER_PROFILE["jacket_sizes_it"]):
        fit = 8
        reasons.append("tailoring size in target")
    elif BUYER_PROFILE["pants_waist"] in sz_text:
        fit = 8
        reasons.append(f"waist in target ({BUYER_PROFILE['pants_waist']})")
    elif BUYER_PROFILE["shirt_neck"] in sz_text or BUYER_PROFILE["body_size"] in sz_text:
        fit = 7
        reasons.append("shirt/body size in target")

    # Price score
    price = detail.get("current_price") or 0
    buy_now = detail.get("buy_now_price") or 0
    visible = buy_now if buy_now else price
    if visible and visible <= 30:
        price_score = 8
    elif visible and visible <= 75:
        price_score = 6
    elif visible and visible > 150:
        price_score = 3
    else:
        price_score = 5

    quality = min(10, 3 + min(len(brands), 3) * 2 + min(len(fabrics), 2))
    rarity = min(10, 3 + len(brands) + len(fabrics))
    taste = min(10, 4 + len(fabrics) + (2 if brands else 0))

    if reject:
        quality = min(quality, 2)
        fit = min(fit, 1)
    if mall:
        quality = min(quality, 4)

    scores = {
        "fit": max(1, min(10, fit)),
        "fabric_quality": max(1, min(10, min(10, 2 + min(len(fabrics), 5) * 2))),
        "maker_quality": max(1, min(10, min(10, 2 + min(len(brands), 4) * 2))),
        "value": max(1, min(10, price_score)),
        "rarity": max(1, min(10, rarity)),
        "taste_fit": max(1, min(10, taste)),
    }

    avg = sum(scores.values()) / len(scores)
    severe = bool(red_flags and any("reject" in f or "paused" in f for f in red_flags))
    if severe or not detail.get("live_verified"):
        recommendation = "Skip"
    elif avg >= 7.0:
        recommendation = "Buy"
    elif avg >= 5.5:
        recommendation = "Watch"
    elif avg >= 4.5:
        recommendation = "Need measurements"
    else:
        recommendation = "Skip"

    if not reasons:
        reasons.append("keyword match only")

    return {
        **detail,
        "query": query,
        "match_reasons": " | ".join(reasons),
        "red_flags": " | ".join(red_flags),
        "total_score": round(avg, 2),
        "recommendation": recommendation,
        **{f"score_{k}": v for k, v in scores.items()},
    }

# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_report(candidates: list[dict], output_dir: Path) -> None:
    top = sorted(
        [c for c in candidates if c["recommendation"] in ("Buy", "Watch") and c["live_verified"]],
        key=lambda c: c["total_score"], reverse=True,
    )[:20]
    maybe = sorted(
        [c for c in candidates if c["recommendation"] == "Need measurements" and c["live_verified"]],
        key=lambda c: c["total_score"], reverse=True,
    )[:20]

    lines = [
        "# ShopTheSalvationArmy Menswear Hunt",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"Candidates verified: {len(candidates)}",
        "",
        "## Top Picks",
    ]
    for i, c in enumerate(top, 1):
        lines += [
            f"### {i}. [{c['title']}]({c['url']})",
            f"- **Item ID:** {c['item_id']}",
            f"- **Search string:** `{c['query']}`",
            f"- **Score:** {c['total_score']}/10 — **{c['recommendation']}**",
            f"- **Price:** current ${c['current_price']} / Buy Now ${c['buy_now_price'] or 0} / shipping ${c['shipping'] or '?'}",
            f"- **Ends:** {c['end_time'] or 'not verified'} — {c['num_bids']} bid(s)",
            f"- **Condition:** {c['condition'] or 'not verified'}",
            f"- **Why:** {c['match_reasons']}",
            f"- **Flags:** {c['red_flags'] or 'none'}",
            f"- **Images:** {'; '.join(c['image_urls'][:2]) or 'none'}",
            "",
        ]
    lines.append("## Maybe / Need Measurements")
    for c in maybe:
        lines.append(f"- [{c['title']}]({c['url']}) — {c['total_score']}/10 | `{c['query']}` | ${c['current_price']} | {c['red_flags'] or 'no flags'}")

    (output_dir / "report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def save_outputs(candidates: list[dict], output_dir: Path) -> None:
    (output_dir / "candidates.json").write_text(json.dumps(candidates, indent=2), encoding="utf-8")
    if candidates:
        with (output_dir / "candidates.csv").open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(candidates[0].keys()))
            writer.writeheader()
            writer.writerows(candidates)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            colors = {"Buy": "C6EFCE", "Watch": "FFEB9C", "Need measurements": "DDEBF7"}
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Candidates"
            headers = list(candidates[0].keys())
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(bold=True, color="FFFFFF")
            for row_idx, row in enumerate(candidates, 2):
                fill = colors.get(row.get("recommendation", ""))
                for col, h in enumerate(headers, 1):
                    v = row[h]
                    if isinstance(v, list):
                        v = " | ".join(str(x) for x in v)
                    cell = ws.cell(row=row_idx, column=col, value=v)
                    if fill:
                        cell.fill = PatternFill("solid", fgColor=fill)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 2, 60
                )
            wb.save(output_dir / "candidates.xlsx")
        except ImportError:
            pass
    write_report(candidates, output_dir)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Category browse (active listings only)
# STSA's keyword search requires JavaScript; we browse all active listings
# in the broad Clothing & Shoes category and filter by brand name client-side.
# ---------------------------------------------------------------------------

BROWSE_CATEGORIES = [
    # (url, label)
    ("https://www.shopthesalvationarmy.com/Browse/C160719/Clothing-Shoes-Accessories"
     "?StatusFilter=active_only&SortFilterOptions=1", "Clothing+Shoes"),
]

# Brand terms to match against listing titles (pre-fetch filter)
BRAND_SIGNAL_WORDS = {
    # Shoes
    "allen edmonds", "alden", "red wing", "viberg", "wesco", "white's",
    "ferragamo", "salvatore", "santoni", "magnanni", "tod's", "tod ",
    "carmina", "carlos santos", "meermin", "testoni", "bontoni", "bally",
    "crockett", "edward green", "john lobb", "gaziano", "church's", "tricker",
    "grenson", "loake", "cheaney", "sargent", "cleverley", "foster",
    "paraboot", "berluti", "weston", "heschung", "corthay",
    "quoddy", "esquivel", "visvim", "vass", "sanders",
    # Tailoring
    "canali", "corneliani", "zegna", "brioni", "kiton", "isaia", "boglioli",
    "sartorio", "caruso", "loro piana", "brunello", "cucinelli",
    "armani", "hickey freeman", "oxxford", "j. press", "southwick", "paul stuart",
    "hackett", "crombie", "crombie", "grenfell", "belstaff", "barbour",
    "barbour", "mackintosh", "aquascutum",
    "cifonelli", "arnys", "husbands", "officine generale", "lemaire",
    "hermes", "hermès", "charvet", "charvet",
    # Shirts
    "turnbull", "asser", "hilditch", "harvie", "lingwood", "emma willis",
    "finamore", "borrelli", "barba", "fray", "lorenzini", "gitman", "kamakura",
    "charvet", "eton",
    # Knitwear
    "drumohr", "inis méain", "inis meain", "john smedley", "smedley",
    "ballantyne", "pringle", "malo",
    # Workwear
    "filson", "vetra", "le laboureur",
    # Blankets
    "pendleton", "faribault", "faribo", "hudson's bay", "hudson bay",
}


def title_has_signal(title: str) -> bool:
    t = title.lower()
    return any(b in t for b in BRAND_SIGNAL_WORDS)


def browse_active_page(session: requests.Session, base_url: str, page: int) -> tuple[list[dict], bool]:
    """Fetch one page of active category browse. Returns (items, has_more)."""
    url = f"{base_url}&page={page}"
    resp = session.get(url, timeout=20)
    resp.raise_for_status()
    items = parse_gallery_page(resp.text)
    # Detect if there's a next page
    has_more = bool(re.search(rf'page={page + 1}', resp.text))
    return items, has_more


def main() -> None:
    parser = argparse.ArgumentParser(description="ShopTheSalvationArmy menswear hunter.")
    parser.add_argument("--output", default="runs/stsa_latest")
    parser.add_argument("--max-pages", type=int, default=90, help="Max category pages to browse")
    parser.add_argument("--max-detail", type=int, default=200)
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    output_dir = Path(args.output)
    if not output_dir.is_absolute():
        output_dir = script_dir / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update(BROWSER_HEADERS)

    seen: set[int] = set()
    candidates: list[dict] = []
    total_browsed = 0
    total_matched = 0

    for cat_url, cat_label in BROWSE_CATEGORIES:
        print(f"Browsing: {cat_label}")
        for page in range(1, args.max_pages + 1):
            try:
                stubs, has_more = browse_active_page(session, cat_url, page)
            except Exception as exc:
                print(f"  page {page} error: {exc}")
                break
            total_browsed += len(stubs)
            # Filter by brand signal in title
            hits = [s for s in stubs if title_has_signal(s.get("title", "")) and not s.get("status_closed")]
            total_matched += len(hits)
            if hits:
                print(f"  page {page}: {len(stubs)} items, {len(hits)} brand matches")
            for stub in hits:
                item_id = stub["item_id"]
                if not item_id or item_id in seen:
                    continue
                seen.add(item_id)
                if len(candidates) >= args.max_detail:
                    break
                title_lower = stub.get("title", "").lower()
                if any(b in title_lower for b in REJECT_BRANDS):
                    continue
                try:
                    detail = fetch_detail(session, item_id)
                    if not detail.get("live_verified"):
                        continue
                    scored = score_candidate(detail, stub.get("title", ""))
                    if not scored["image_urls"] and stub.get("image_url"):
                        scored["image_urls"] = [stub["image_url"]]
                    candidates.append(scored)
                except Exception as exc:
                    print(f"  detail error {item_id}: {exc}")
                time.sleep(0.25)

            if len(candidates) >= args.max_detail:
                break
            if not has_more:
                print(f"  reached last page ({page})")
                break
            time.sleep(0.3)

    print(f"\nBrowsed {total_browsed} active listings, {total_matched} brand matches")
    candidates.sort(key=lambda c: c["total_score"], reverse=True)
    save_outputs(candidates, output_dir)
    print(f"Done. {len(candidates)} candidates → {output_dir}")


if __name__ == "__main__":
    main()
